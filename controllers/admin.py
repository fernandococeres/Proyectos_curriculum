"""
Módulo administrativo para gestión avanzada del Sistema de Gestión Escolar. Contiene funcionalidades para análisis de datos, generación de estadísticas y exportación de información en formatos CSV y Excel. Este módulo está diseñado para ser utilizado por administradores y personal con permisos avanzados, proporcionando herramientas para la toma de decisiones informadas y la gestión eficiente de la información escolar.
"""
from models.dbConn import dbConn
from datetime import datetime
import csv
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False


class AdminAlumnos:
    """Administrador avanzado para alumnos."""
    
    def __init__(self, db_name="escuela.db"):
        self.db = dbConn(db_name)
    
    # ===== ESTADÍSTICAS =====
    def get_estadisticas_generales(self) -> dict:
        """Obtiene estadísticas generales de alumnos."""
        alumnos = self.db.execute("SELECT * FROM alumnos")
        
        return {
            "total_alumnos": len(alumnos),
            "grados": self._contar_por_grado(alumnos),
            "secciones": self._contar_por_seccion(alumnos),
            "alumnos_por_grado_seccion": self._agrupar_grado_seccion(alumnos)
        }
    
    def _contar_por_grado(self, alumnos: list) -> dict:
        """Cuenta alumnos por grado."""
        contador = {}
        for alumno in alumnos:
            grado = alumno[6]  # índice de grado
            contador[grado] = contador.get(grado, 0) + 1
        return contador
    
    def _contar_por_seccion(self, alumnos: list) -> dict:
        """Cuenta alumnos por sección."""
        contador = {}
        for alumno in alumnos:
            seccion = alumno[7]  # índice de sección
            contador[seccion] = contador.get(seccion, 0) + 1
        return contador
    
    def _agrupar_grado_seccion(self, alumnos: list) -> dict:
        """Agrupa alumnos por grado y sección."""
        agrupado = {}
        for alumno in alumnos:
            grado = alumno[6]
            seccion = alumno[7]
            key = f"{grado} {seccion}"
            agrupado[key] = agrupado.get(key, 0) + 1
        return dict(sorted(agrupado.items()))
    
    def get_alumnos_por_grado(self, grado: str) -> list:
        """Obtiene todos los alumnos de un grado específico."""
        comando = "SELECT * FROM alumnos WHERE grado = ? ORDER BY apellido, nombre"
        return self.db.execute(comando, (grado,))
    
    def get_alumnos_por_seccion(self, seccion: str) -> list:
        """Obtiene todos los alumnos de una sección específica."""
        comando = "SELECT * FROM alumnos WHERE seccion = ? ORDER BY grado, apellido, nombre"
        return self.db.execute(comando, (seccion,))
    
    def get_alumnos_sin_email(self) -> list:
        """Obtiene alumnos sin email registrado."""
        comando = "SELECT * FROM alumnos WHERE email IS NULL OR email = '' ORDER BY apellido, nombre"
        return self.db.execute(comando)
    
    def get_alumnos_recientes(self, dias: int = 7) -> list:
        """Obtiene alumnos registrados en los últimos N días."""
        comando = """SELECT * FROM alumnos 
                    WHERE datetime(fecha_registro) >= datetime('now', '-' || ? || ' days')
                    ORDER BY fecha_registro DESC"""
        return self.db.execute(comando, (dias,))
    
    # ===== EXPORTACIÓN =====
    def exportar_a_csv(self, archivo: str = None) -> str:
        """Exporta todos los alumnos a CSV."""
        if not archivo:
            archivo = f"alumnos_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        alumnos = self.db.execute("SELECT * FROM alumnos ORDER BY apellido, nombre")
        
        try:
            with open(archivo, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')
                # Header
                writer.writerow(['ID', 'Nombre', 'Apellido', 'Email', 'Teléfono', 
                               'Fecha Nacimiento', 'Grado', 'Sección', 'DNI', 'Fecha Registro'])
                # Datos
                for alumno in alumnos:
                    writer.writerow(alumno)
            
            return archivo
        except Exception as e:
            raise Exception(f"Error al exportar CSV: {e}")
    
    def exportar_a_excel(self, archivo: str = None) -> str:
        """Exporta todos los alumnos a Excel."""
        if not EXCEL_DISPONIBLE:
            raise Exception("openpyxl no está instalado. Instala: pip install openpyxl")
        
        if not archivo:
            archivo = f"alumnos_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        alumnos = self.db.execute("SELECT * FROM alumnos ORDER BY apellido, nombre")
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Alumnos"
            
            # Encabezados
            headers = ['ID', 'Nombre', 'Apellido', 'Email', 'Teléfono', 
                      'Fecha Nacimiento', 'Grado', 'Sección', 'DNI', 'Fecha Registro']
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Datos
            for row_idx, alumno in enumerate(alumnos, 2):
                for col_idx, value in enumerate(alumno, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 8
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 10
            ws.column_dimensions['J'].width = 20
            
            # Hoja de estadísticas
            self._agregar_estadisticas_excel(wb)
            
            wb.save(archivo)
            return archivo
        except Exception as e:
            raise Exception(f"Error al exportar Excel: {e}")
    
    def _agregar_estadisticas_excel(self, wb):
        """Agrega una hoja con estadísticas al workbook."""
        stats = self.get_estadisticas_generales()
        
        ws_stats = wb.create_sheet("Estadísticas")
        
        # Título
        ws_stats['A1'] = "ESTADÍSTICAS DE ALUMNOS"
        ws_stats['A1'].font = Font(bold=True, size=14)
        ws_stats['A1'].alignment = Alignment(horizontal="center")
        
        # Total
        ws_stats['A3'] = "Total de Alumnos:"
        ws_stats['B3'] = stats['total_alumnos']
        
        # Por grado
        ws_stats['A5'] = "Alumnos por Grado"
        ws_stats['A5'].font = Font(bold=True)
        row = 6
        for grado, cantidad in sorted(stats['grados'].items()):
            ws_stats[f'A{row}'] = f"Grado {grado}"
            ws_stats[f'B{row}'] = cantidad
            row += 1
        
        # Por sección
        ws_stats['D5'] = "Alumnos por Sección"
        ws_stats['D5'].font = Font(bold=True)
        row = 6
        for seccion, cantidad in sorted(stats['secciones'].items()):
            ws_stats[f'D{row}'] = f"Sección {seccion}"
            ws_stats[f'E{row}'] = cantidad
            row += 1
        
        # Ajustar anchos
        ws_stats.column_dimensions['A'].width = 20
        ws_stats.column_dimensions['B'].width = 15
        ws_stats.column_dimensions['D'].width = 20
        ws_stats.column_dimensions['E'].width = 15


class AdminProfesores:
    """Administrador avanzado para profesores."""
    
    def __init__(self, db_name="escuela.db"):
        self.db = dbConn(db_name)
    
    # ===== ESTADÍSTICAS =====
    def get_estadisticas_generales(self) -> dict:
        """Obtiene estadísticas generales de profesores."""
        profesores = self.db.execute("SELECT * FROM profesores")
        
        return {
            "total_profesores": len(profesores),
            "especialidades": self._contar_por_especialidad(profesores),
            "profesores_sin_email": len([p for p in profesores if not p[3]])
        }
    
    def _contar_por_especialidad(self, profesores: list) -> dict:
        """Cuenta profesores por especialidad."""
        contador = {}
        for profesor in profesores:
            especialidad = profesor[5]  # índice de especialidad
            contador[especialidad] = contador.get(especialidad, 0) + 1
        return dict(sorted(contador.items()))
    
    def get_profesores_por_especialidad(self, especialidad: str) -> list:
        """Obtiene profesores de una especialidad específica."""
        comando = "SELECT * FROM profesores WHERE especialidad = ? ORDER BY apellido, nombre"
        return self.db.execute(comando, (especialidad,))
    
    def get_especialidades(self) -> list:
        """Obtiene lista de todas las especialidades."""
        especialidades = self.db.execute("SELECT DISTINCT especialidad FROM profesores ORDER BY especialidad")
        return [e[0] for e in especialidades if e[0]]
    
    def get_profesores_sin_email(self) -> list:
        """Obtiene profesores sin email registrado."""
        comando = "SELECT * FROM profesores WHERE email IS NULL OR email = '' ORDER BY apellido, nombre"
        return self.db.execute(comando)
    
    def get_profesores_sin_documento(self) -> list:
        """Obtiene profesores sin documento registrado."""
        comando = "SELECT * FROM profesores WHERE documento IS NULL OR documento = '' ORDER BY apellido, nombre"
        return self.db.execute(comando)
    
    def get_profesores_recientes(self, dias: int = 7) -> list:
        """Obtiene profesores registrados en los últimos N días."""
        comando = """SELECT * FROM profesores 
                    WHERE datetime(fecha_registro) >= datetime('now', '-' || ? || ' days')
                    ORDER BY fecha_registro DESC"""
        return self.db.execute(comando, (dias,))
    
    # ===== EXPORTACIÓN =====
    def exportar_a_csv(self, archivo: str = None) -> str:
        """Exporta todos los profesores a CSV."""
        if not archivo:
            archivo = f"profesores_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        profesores = self.db.execute("SELECT * FROM profesores ORDER BY apellido, nombre")
        
        try:
            with open(archivo, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')
                # Header
                writer.writerow(['ID', 'Nombre', 'Apellido', 'Email', 'Teléfono', 
                               'Especialidad', 'Documento', 'Fecha Registro'])
                # Datos
                for profesor in profesores:
                    writer.writerow(profesor)
            
            return archivo
        except Exception as e:
            raise Exception(f"Error al exportar CSV: {e}")
    
    def exportar_a_excel(self, archivo: str = None) -> str:
        """Exporta todos los profesores a Excel."""
        if not EXCEL_DISPONIBLE:
            raise Exception("openpyxl no está instalado. Instala: pip install openpyxl")
        
        if not archivo:
            archivo = f"profesores_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        profesores = self.db.execute("SELECT * FROM profesores ORDER BY apellido, nombre")
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Profesores"
            
            # Encabezados
            headers = ['ID', 'Nombre', 'Apellido', 'Email', 'Teléfono', 
                      'Especialidad', 'Documento', 'Fecha Registro']
            
            header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Datos
            for row_idx, profesor in enumerate(profesores, 2):
                for col_idx, value in enumerate(profesor, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 20
            
            # Hoja de estadísticas
            self._agregar_estadisticas_excel(wb)
            
            wb.save(archivo)
            return archivo
        except Exception as e:
            raise Exception(f"Error al exportar Excel: {e}")
    
    def _agregar_estadisticas_excel(self, wb):
        """Agrega una hoja con estadísticas al workbook."""
        stats = self.get_estadisticas_generales()
        
        ws_stats = wb.create_sheet("Estadísticas")
        
        # Título
        ws_stats['A1'] = "ESTADÍSTICAS DE PROFESORES"
        ws_stats['A1'].font = Font(bold=True, size=14)
        ws_stats['A1'].alignment = Alignment(horizontal="center")
        
        # Total
        ws_stats['A3'] = "Total de Profesores:"
        ws_stats['B3'] = stats['total_profesores']
        
        # Sin email
        ws_stats['A4'] = "Profesores sin Email:"
        ws_stats['B4'] = stats['profesores_sin_email']
        
        # Por especialidad
        ws_stats['A6'] = "Profesores por Especialidad"
        ws_stats['A6'].font = Font(bold=True)
        row = 7
        for especialidad, cantidad in stats['especialidades'].items():
            ws_stats[f'A{row}'] = especialidad
            ws_stats[f'B{row}'] = cantidad
            row += 1
        
        # Ajustar anchos
        ws_stats.column_dimensions['A'].width = 25
        ws_stats.column_dimensions['B'].width = 15
